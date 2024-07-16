# Create your views here.
import datetime
from io import BytesIO
import io
import json
import time
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.urls import reverse, reverse_lazy
from openpyxl import Workbook,load_workbook
from redis import Redis

from clinicals.models import Patient
from .forms import FileUploadForm
from .models import UploadedFile
from PyPDF2 import PdfMerger, PdfReader
from django.core.cache import cache
from openpyxl.styles import PatternFill
import os
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.core.paginator import EmptyPage,PageNotAnInteger,Paginator
from django.views.generic import CreateView,DeleteView,UpdateView,ListView
from django.core.files import File


def color_header(sheet,color):
    header_row = sheet[1] 
    for cell in header_row:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
def mergeMultipleExcel(request):
    media_files_cache_key='media_files'
    get_cache_data=cache.get(media_files_cache_key)
    
    if get_cache_data:
            media_data=get_cache_data
            print("This list data are from the cache memory")  
    else:
        
        media_data = UploadedFile.objects.all()
        cache.set(media_files_cache_key,media_data,timeout=100)

    if request.method == "POST":
        files = request.POST.getlist("selected_files")
        if not files:
            error_message = "No files selected. Please select at least one file."
            return HttpResponse(error_message, status=400) 
        output_wb = Workbook()
        output_ws = output_wb.active 
        output_pdf = PdfMerger()
        output_stream = BytesIO()

        for uploaded_file_id in files:
            uploaded_file = get_object_or_404(UploadedFile, pk=uploaded_file_id)
            file_path = uploaded_file.file.path
            try:
                if file_path.endswith('.xlsx'):
                    wb = load_workbook(file_path)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            output_ws.append(row)
                            # Color the header row
                            if row_index == 1:
                                color_header(output_ws, "D3D3D3") 
                    
                    output_stream = BytesIO()
                    output_wb.save(output_stream) 
                    output_stream.seek(0)
        
                elif file_path.endswith('.pdf'):
                    output_wb = Workbook()
                    output_ws = output_wb.active 
                    pdf_reader = PdfReader(file_path)
                    for page_num in range(len(pdf_reader.pages)):
                        page = pdf_reader.pages[page_num]
                        for line in page.extract_text().split('\n'):
                            output_ws.append([line])
                    output_pdf.append(file_path)
                    
                    output_stream = BytesIO()
                    output_pdf.write(output_stream)
                    output_stream.seek(0)
                else:
                    return HttpResponse("Error: Unsupported file format")

            except Exception as e:
                return HttpResponse(f"Error: {e}")

        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' if any(file_path.endswith('.xlsx') for file_path in [uploaded_file.file.path for uploaded_file in UploadedFile.objects.filter(pk__in=files)]) else 'application/pdf'
        filename = 'merged.xlsx' if content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' else 'merged.pdf'
        response = HttpResponse(output_stream, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response 
    
# def new_media_data():
#     redis_connection=Redis(host='redis', port=6379, db=1)
#     key="media_files_list"
#     get_redis_data=redis_connection.lrange(key,0, -1)
#     print("I am am executed now ")
#     media_data = []
#     for item in get_redis_data:
#         decoded_item = item.decode('utf-8')
#         media_data.append(json.loads(decoded_item))
#     redis_connection.set("media_files_all", json.dumps(media_data)) 
    

def mergeMultipleExcelInMultipleSheet(request, update_instance=None):
    print("I am from merge excel")
    if update_instance:
        print("I am also executed")
    redis_connection=Redis(host='redis', port=6379, db=1)
    inital_page = request.GET.get('page', 1)
    items_per_page = 20
    try:
        page = int(inital_page)
    except ValueError:
        page = 1

    start_index = (page - 1) * items_per_page
    end_index = start_index + items_per_page - 1
    
    start_redis_time = time.time()
    redis_data_chunk = redis_connection.lrange("media_files_list", start_index, end_index)
    total_items = redis_connection.llen("media_files_list") # used to calculate the pages 
    end_redis_time = time.time()
    redis_time_taken = end_redis_time - start_redis_time
    print("Time taken to fetch the data from Redis memory: ", redis_time_taken)
    
    if redis_data_chunk:
        start = time.time()
        redis_data = [json.loads(item) for item in redis_data_chunk]
        end = time.time()
        print("Time to decode the redis_data: ", end - start)
    else:
        print("Data is fetching from the DB ! ")
        media_data=list(UploadedFile.objects.all().values().order_by('-id'))

        if media_data:
            start=time.time()
            for item in media_data:
                redis_connection.rpush("media_files_list", json.dumps(item))
            end=time.time()
            print("Time Taken to store the 9 lakh data in the Redis from Database",end-start)
            
        redis_data_chunk = redis_connection.lrange("media_files_list", start_index, end_index)
        redis_data = [json.loads(item) for item in redis_data_chunk]
        total_items = redis_connection.llen("media_files_list")
        
    paginator = Paginator(range(total_items), items_per_page)
        
    try:
        objects = paginator.page(page)
    except PageNotAnInteger:
        objects = paginator.page(1)
           
    paginator_obj = objects  # this contains the object of the pages
    ui_data=redis_data # this contains the actual JSON data     
    print(ui_data)

    if request.method == "POST":
        if not request.POST.get('update') and not request.POST.get('delete'):
            print("This is the request data ",request)
            files = request.POST.getlist("selected_files")
            print(files)
            if not files:
                error_message = "No files selected. Please select at least one file."
                return HttpResponse(error_message, status=400) 
            
            output_wb = Workbook()
            output_wb.remove(output_wb.active)
            
            output_pdf = PdfMerger()
            output_stream = BytesIO()
            
            # checking if user sends multiple files like xlsx and pdf at same time
            file_count=0
            xls_files=0
            pdf_files=0
                
            for selectedfile_id in files:
                selectedfile = get_object_or_404(UploadedFile, pk=selectedfile_id)
                file_path = selectedfile.file.path
                
                file_count=file_count+1
                if file_path.endswith('.xlsx'):
                    xls_files=xls_files+1
                if file_path.endswith('.pdf'):
                        pdf_files=pdf_files+1
                        
            print("total Files uploaded",file_count)
            print("total xls Files found ",xls_files)
            print("total pdf Files uploaded",pdf_files)


            for uploaded_file_id in files:
                uploaded_file = get_object_or_404(UploadedFile, pk=uploaded_file_id)
                file_path = uploaded_file.file.path
                try:
                    if file_count==xls_files:
                        
                        file_name_with_extension = os.path.basename(uploaded_file.file.name)
                        file_name_without_extension = os.path.splitext(file_name_with_extension)[0]
                        
                        wb = load_workbook(file_path)
                        for sheet_name in wb.sheetnames:
                            output_ws = output_wb.create_sheet(title=file_name_without_extension)
                            sheet = wb[sheet_name]
                            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                                output_ws.append(row)
                                if row_index == 1:
                                    color_header(output_ws, "D3D3D3") 
                        
                        output_stream = BytesIO()
                        output_wb.save(output_stream) 
                        output_stream.seek(0)
            
                    elif file_count==pdf_files:
                        output_wb = Workbook()
                        output_ws = output_wb.active 
                        pdf_reader = PdfReader(file_path)
                        for page_num in range(len(pdf_reader.pages)):
                            page = pdf_reader.pages[page_num]
                            for line in page.extract_text().split('\n'):
                                output_ws.append([line])
                        output_pdf.append(file_path)
                        
                        output_stream = BytesIO()
                        output_pdf.write(output_stream)
                        output_stream.seek(0)
                    else:
                        return HttpResponse("Error: Files are of multiple format!.")

                except Exception as e:
                    return HttpResponse(f"Error: {e}")
        elif request.POST.get('update') :
            current_page=int(request.POST.get('page_number'))
            id=int(request.POST.get('selected_files'))        
            
            start_index = (current_page - 1) * items_per_page
            end_index = start_index + items_per_page - 1 

            redis_data_chunk = redis_connection.lrange("media_files_list", start_index, end_index)
            new_redis_data = [json.loads(item) for item in redis_data_chunk]

            for item in new_redis_data:
                if id == int(item.get('id')):    
                    storeDataToUpdate(start_index,end_index,id)
                    url = reverse('updateTest', kwargs={'pk': id})
                    return redirect(url)

        elif request.POST.get('delete'):
            current_page=int(request.POST.get('page_number'))
            id=int(request.POST.get('selected_files'))
        
            start_index = (current_page - 1) * items_per_page
            end_index = start_index + items_per_page - 1 
            
            redis_data_chunk = redis_connection.lrange("media_files_list", start_index, end_index)
            new_redis_data = [json.loads(item) for item in redis_data_chunk]
            
            for item in new_redis_data:
                if id == int(item.get('id')):
                    item_to_remove = json.dumps(item)
                    redis_connection.lrem("media_files_list", 0, item_to_remove)
                    break
    
            url = reverse('deleteData', kwargs={'pk': id})
            return redirect(url)
            

        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' if any(file_path.endswith('.xlsx') for file_path in [uploaded_file.file.path for uploaded_file in UploadedFile.objects.filter(pk__in=files)]) else 'application/pdf'
        filename = 'merged.xlsx' if content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' else 'merged.pdf'
        response = HttpResponse(output_stream, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response 
    

    return render(request, "mergeMultipleExcel/upload.html",{'pages':paginator_obj,'ui_data':ui_data})

def storeDataToUpdate(si=None,ei=None,data_id=None,update_instance=None):
    r = Redis(host='redis', port=6379, db=1)
    if si is not None and ei is not None and data_id is not None:
        r.set('start_index', si)
        r.set('end_index', ei)
        r.set('data_id', data_id)
        
    if update_instance is not None:        
        start_index= int(r.get('start_index'))
        end_index= int(r.get('end_index'))
        id=int(r.get("data_id"))
        
        redis_data_chunk = r.lrange("media_files_list", start_index, end_index)
        redis_data = [json.loads(item) for item in redis_data_chunk]
        
        for item in redis_data:
            if id == int(item.get('id')):
                item.update(serialize_instance(update_instance))
        
        # Serialize the updated chunk into JSON strings
        updated_redis_data_chunk = [json.dumps(item) for item in redis_data]
        for index, item_json in enumerate(updated_redis_data_chunk):
            # The list indices are assumed to start from start_index.
            r.lset("media_files_list", start_index + index, item_json)

def serialize_instance(instance):
    data = {}
    for field in instance._meta.fields:
        value = getattr(instance, field.name)
        if isinstance(value, File):
            # Serialize file URL
            data[field.name] = value.url if value else None
        elif isinstance(value, (int, float, str, bool, list, dict)):
            # Handle basic data types
            data[field.name] = value
        else:
            # For unsupported data types, convert to string or handle as needed
            data[field.name] = str(value)
    return data


def multipleProcedureMultipleSheet(request):
    wb=Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    #example json data/ query data
    
    procedureData1 = [{"name":"Andro","contactNo":54532666,"employeeId":98765,"EmployeeType":"KeysightFTE"},
                    {"name":"Sachin","contactNo":54532666,"employeeId":98765,"EmployeeType":"IT Manager"},
                    {"name":"Rolisha","contactNo":54532666,"employeeId":98765,"EmployeeType":"KeysightFTE"}]
    
    procedureData2 = [{"name":"Manish","contactNo":54532666,"employeeId":98765,"EmployeeType":"IT Manager"},
                    {"name":"Ronick","contactNo":54532666,"employeeId":98765,"EmployeeType":"KeysightFTE"},
                    {"name":"Kamuna","contactNo":54532666,"employeeId":98765,"EmployeeType":"IT Manager"}]
    
    all_json_data = [procedureData1, procedureData2]
    print(all_json_data)
    if request.method == "POST":
        for index, json_data in enumerate(all_json_data, start=1):
            sheet = wb.create_sheet(title=f"Sheet {index}")
            # Write column headers
            for col, key in enumerate(json_data[0].keys(), start=1):
                cell=sheet.cell(row=1, column=col, value=key)
                cell.fill = header_fill
            # Write data
            for row, emp in enumerate(json_data, start=2):
                for col, value in enumerate(emp.values(), start=1):
                    sheet.cell(row=row, column=col, value=value)
                    

        output = io.BytesIO()
        wb.save(output)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=output.xlsx'
        output.seek(0)
        response.write(output.getvalue())
        
        return response
    
    return render(request, "multipleProcedureMultipleSheet/getsheet.html",{'all_json_data':all_json_data})


# def upload_file(request):
#     if request.method == 'POST':
#         form = FileUploadForm(request.POST, request.FILES)
#         uploaded_instances = []
#         if form.is_valid():
#             for _ in range(200000): 
#                 uploaded_data = form.save(commit=False)
#                 uploaded_instances.append(uploaded_data)
#                 # delete_redis_data()
#             UploadedFile.objects.bulk_create(uploaded_instances)
#             return redirect('upload_success')
#     else:
#         form = FileUploadForm()
#     return render(request, 'upload_form.html', {'form': form})

def upload_file(request):
    redis_connection = Redis(host='redis', port=6379, db=1)
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_data = form.save()
            update_redis_data(uploaded_data,redis_connection)  
            return redirect('upload_success')  
    else:
        form = FileUploadForm()
    return render(request, 'upload_form.html', {'form': form})


class RedisUpdateView(UpdateView):
    model=UploadedFile
    success_url=reverse_lazy('upload_success')
    fields={'file_name', 
            'created_by',
            'file_type',
            'file_size',
            'file',
            'description',
            'is_public',
            'file_extension',
            'tags',
            'is_archived',
            'category'}
    
    def form_valid(self, form):
        response = super().form_valid(form)
        updated_instance = form.instance
        print(updated_instance)
        storeDataToUpdate(update_instance=updated_instance)
        return response
    
class RedisDeleteView(DeleteView):
    model=UploadedFile
    success_url=reverse_lazy('delete_success')
    
# def update_data(request,id):
#     print("items received : ", id)
#     print("Form is being executed ")
#     uploaded_file = get_object_or_404(UploadedFile, id=id)
#     # redis_connection = Redis(host='redis', port=6379, db=1)
#     if request.method == 'POST':
#         form = UpdateForm(request.POST, instance=uploaded_file)
#         if form.is_valid():
#             form.save()
#             # uploaded_data = form.save()
#             # update_redis_data(uploaded_data,redis_connection)  
#             return redirect('upload_success')  
#     else:
#         print("items received : ", uploaded_file)
#         form = UpdateForm(instance=uploaded_file)
#     return render(request, 'upload_form.html', {'form': form})

def delete_redis_data():
    media_files_cache_key = 'media_files_all'
    redis_connection = Redis(host='redis', port=6379, db=1)
    redis_connection.delete(media_files_cache_key)  

def update_redis_data(uploaded_data,redis_connection):
    start=time.time() 
    new_data={
        'id': uploaded_data.id,
        'file_name': uploaded_data.file_name,
        'created_by': uploaded_data.created_by,
        'file_type': uploaded_data.file_type,
        'file_size': uploaded_data.file_size,
        'description' :uploaded_data.description,
        'is_public' :uploaded_data.is_public,
        'file_extension' :uploaded_data.file_extension,
        'tags' : uploaded_data.tags,
        'is_archived' :uploaded_data.is_archived,
        'category' :uploaded_data.category,
    }
    redis_connection.lpush("media_files_list", json.dumps(new_data))
    end=time.time()
    print("time Taken to update the List is ",end-start)
    
def upload_success(request):
    return render(request, 'upload_success.html')

def delete_success(request):
    return render(request, 'delete_success.html')

def show_uploaded_data(request):
    all_data=UploadedFile.objects.all()
    return render(request,'mergeMultipleExcel/upload.html',{'data':all_data})

# http://localhost:8080/get_no_files/
#use this url to call this api 

@api_view(['GET'])
def get_no_files(request):
    start=time.time()
    redis_key='total_number_of_files'
    redis_connection = Redis(host='redis', port=6379, db=1)
    get_redis_data=redis_connection.get(redis_key)
    
    if get_redis_data:
        redis_connection.delete(redis_key)  
        start_redis_time=time.time()
        total_files = json.loads(get_redis_data.decode('utf-8'))
        print("This data is being fetched from the redis-cache memory")
            
        end_redis_time = time.time()
        redis_time_taken = end_redis_time - start_redis_time
        print("Time taken to fetch the data from the redis-memory ", redis_time_taken)
        print("json dumps value ",json.dumps(total_files))
        return Response({'count': total_files,'redis time_taken':redis_time_taken})
    else:
        data=UploadedFile.objects.all()
        count = 0
        if data:
            for item in data:
                count = count + 1 
        end=time.time()
        redis_connection.set(redis_key, json.dumps(count))
        print("json dumps value ",json.dumps(count))
    return Response({'count': count, 'db time_taken': end - start})


@api_view(['GET'])
def getIsPublic(request):
    start=time.time()
    print("getisPublic  API got triggered ! ")
    redis_connection = Redis(host='redis', port=6379, db=1)
    get_is_public_data=redis_connection.get("is_public_data")
    
    filter_data=[]
    
    if get_is_public_data:
        filter_data=json.loads(get_is_public_data)
    else:
        json_data = redis_connection.get("media_files_all")
        
        if json_data:
            data_list = json.loads(json_data)            
            filter_data = [item for item in data_list if item.get('is_public', False)]
        redis_connection.set("is_public_data",json.dumps(filter_data))

    
    end=time.time()
    total_time=end-start
    return Response({'is_Public_data':filter_data,'cache_time':total_time})